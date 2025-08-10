import os
from const import EXCEL_FILE
from openpyxl import load_workbook

# Предполагаем, что файл находится в той же директории, что и этот скрипт,
# или передайте путь явно
DEFAULT_EXCEL_FILENAME = EXCEL_FILE

class TrainScheduleEntry:
    """
    Класс для представления одной строки из Excel-файла расписания.
    Предполагается, что Excel имеет 15 столбцов (A-O).
    """
    def __init__(self, data_list):
        """
        Инициализирует экземпляр данными из списка.
        :param data_list: Список значений из 15 ячеек строки.
        """
        if len(data_list) != 15:
            raise ValueError("Список данных должен содержать ровно 15 элементов.")

        # Предполагаем, что столбцы идут от A до O
        self.col_A = data_list[0]
        self.col_B = data_list[1]
        self.col_C = data_list[2]
        self.col_D = data_list[3]
        self.col_E = data_list[4]
        self.col_F = data_list[5]
        self.col_G = data_list[6]
        self.col_H = data_list[7]
        self.col_I = data_list[8]
        self.col_J = data_list[9]
        self.col_K = data_list[10]
        self.col_L = data_list[11]
        self.col_M = data_list[12]
        self.col_N = data_list[13]
        self.col_O = data_list[14] # Это поле будет проверяться на заполненность

    def __repr__(self):
        # Упрощенное представление для отладки
        return f"TrainScheduleEntry(Route: {self.col_A}, Date (O): {self.col_O})"

    # Пример метода для получения значения для поля ввода (столбец A)
    def get_route_name(self):
        """Возвращает значение из столбца A, предположительно - название маршрута."""
        return str(self.col_A) if self.col_A is not None else ""

    # Добавьте другие методы для доступа к данным по необходимости


def parse_excel_schedule(file_path=None):
    """
    Парсит Excel-файл и возвращает список экземпляров TrainScheduleEntry
    для строк, где столбец A (1-й) и столбец O (15-й) заполнены.
    Использует openpyxl.

    :param file_path: Путь к Excel-файлу. Если None, используется DEFAULT_EXCEL_FILENAME
                      в текущей директории.
    :return: Список экземпляров TrainScheduleEntry.
    """
    if file_path is None:
        # Получаем путь к директории, где лежит этот скрипт
        script_dir = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(script_dir, DEFAULT_EXCEL_FILENAME)

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel-файл не найден: {file_path}")

    entries = []
    try:
        # Открываем книгу с помощью openpyxl
        # data_only=True чтобы получить значения, а не формулы
        wb = load_workbook(filename=file_path, data_only=True)
        # Работаем с активным листом. Если нужно указать конкретный лист,
        # используйте wb["ИмяЛиста"]
        ws = wb.active

        # Итерируемся по строкам, начиная с первой.
        # enumerate(ws.iter_rows(values_only=True), 2) даст номер строки (начиная с 2)
        # и кортеж значений ячеек для каждой строки.
        for row_num, row_tuple in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            
            # Преобразуем кортеж в список и убедимся, что он имеет длину 15
            if len(row_tuple) < 15:
                # Если в строке меньше 15 столбцов, дополняем None
                row_data = list(row_tuple) + [None] * (15 - len(row_tuple))
                # print(f"⚠️  Строка {row_num} имеет менее 15 столбцов, дополнена None.")
            elif len(row_tuple) > 15:
                # Если больше 15, берем только первые 15
                row_data = list(row_tuple[0:15])
                # print(f"⚠️  Строка {row_num} имеет более 15 столбцов, взяты первые 15.")
            else:
                row_data = list(row_tuple)

            # Проверяем 1-й столбец (индекс 0) и 15-й столбец (индекс 14) на заполненность
            value_a = row_data[0]  # Столбец A
            value_o = row_data[14] # Столбец O

            # --- Логика проверки на заполненность ---
            # Проверяем, что оба значения не None и не пустые строки
            # Можно адаптировать под конкретные типы данных (например, isinstance(..., datetime))
            if (value_a is not None and str(value_a).strip() != '' and
                value_o is not None and str(value_o).strip() != ''):
                try:
                    entry = TrainScheduleEntry(row_data)
                    entries.append(entry)
                except ValueError as e:
                    print(f"⚠️  Ошибка создания записи для строки {row_num}: {e}")
            # else:
            #     print(f"ℹ️  Строка {row_num} пропущена: столбец A или O пуст.")

    except Exception as e:
        print(f"❌ Ошибка при парсинге Excel-файла {file_path} (openpyxl): {e}")
        raise # Перебрасываем исключение, чтобы вызывающая сторона могла его обработать

    print(f"✅ Успешно распаршено {len(entries)} записей из Excel (openpyxl, A и O заполнены).")
    return entries